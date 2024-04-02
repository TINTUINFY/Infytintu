import openpyxl, xlsxwriter, pandas as pd
from openpyxl.styles import Alignment
from math import ceil
class ExcelWrapper():
    


    def __init__(self, filename=None):
        if filename:
            self.wb = openpyxl.load_workbook(filename)
            self.sheet = self.wb.active
        
    def get_first_cell(self):
        startlimit = 6
        start = 1
        while(startlimit<1000):
            for r in range(start, startlimit):
                for c in range(start,startlimit):
                    if(self.sheet.cell(row=r, column=c).value != "" and self.sheet.cell(row=r, column=c).value is not None):
                        return (r,c)
                     
            start=startlimit
            startlimit += 5
        return (-1,-1)

    def data_to_excel(self, report, filename, index, sheet, column_width):
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')
        report.to_excel(writer, index=False)
        wb  = writer.book
        sheet = writer.sheets[sheet]
        # Creating the cell formats
        avg_format = wb.add_format({'num_format': '#,##,##0'})
        percent_format = wb.add_format({'num_format':'0.0%'})
        sheet.freeze_panes(1,0)
        # Setting the column format 
        sheet.set_column("A:A",column_width[0])
        sheet.set_column("B:B",column_width[1])
        sheet.set_column("C:C",column_width[2], avg_format)
        sheet.set_column("D:D",column_width[3], avg_format)
        sheet.set_column("E:E",column_width[4], percent_format)
        sheet.set_column("F:F",column_width[5])
        # Adding border to all the cells
        cells = wb.add_format({"border":1})
        sheet.conditional_format(0,0,index,5, {"type": "no_blanks", "format":cells})
        sheet.conditional_format(0,0,index,5, {"type": "blanks", "format":cells})
        writer.save()
        
    def data_to_json(self, report, filename):
        report.to_json(filename)

    def merge_cells(self, filename, total_row_col, merge_row_col, sheetname):
        ws = self.wb[sheetname]
        cells_with_strings = []
        for i in range (1,ws.max_row+1):
            if ws[merge_row_col + str(i)].value != None:
                cells_with_strings.append(int(i))
        merge_cell_range = len(cells_with_strings) - 1
        for i in range (0, (merge_cell_range)):
            ws.merge_cells(merge_row_col + str(cells_with_strings[i]) + ":" + merge_row_col + str((cells_with_strings[(i+1)])-1))

        final_merge = []
        for i in range ((cells_with_strings[-1]), ((cells_with_strings[-1]) + ws.max_row)):
            if ws[total_row_col + str(i)].value != None:
                final_merge.append(int(i))
        ws.merge_cells(merge_row_col + str(final_merge[0]) + ":" + merge_row_col + str(final_merge[-1]))

        for row in ws.iter_rows():
            for cell in row:   
                cell.alignment = Alignment(wrap_text=True,vertical='top')
        ws.row_dimensions[2].height = 50
        for i in range(1, ws.max_row+1):
            ws.row_dimensions[i].height = self.get_height_for_row(ws, i-1, merge_row_col)
        self.wb.save(filename)


    def get_height_for_row(self, sheet, row_number, column_letter):
        factor_of_font_size_to_width = {format: {"factor": 1,"height": 21}}
        font_params = factor_of_font_size_to_width[format]
        row = list(sheet.rows)[row_number]
        height = font_params["height"]

        for cell in row:
            words_count_at_one_row = sheet.column_dimensions[column_letter].width / font_params["factor"]
            lines = ceil(len(str(cell.value)) / words_count_at_one_row)
            height = max(height, lines * font_params["height"])

        return height
